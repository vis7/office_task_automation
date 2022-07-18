"""
This program take certificate image and student details in xlsx file and create
certificates for them based on student details.
"""
###############################################################
# Adding absolute path to the root directory for using absolute path
###############################################################
import os
import sys

abs_path_of_directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.sys.path.append(abs_path_of_directory)
################################################################
import os
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from letter_pdf.utils import check_email, send_mail, create_session, destroy_session

def send_certificate(filepath):
    workbook = load_workbook(filepath) # 'data/student_details.xlsx'
    sheet = workbook.active

    size = 35 
    color = (0,0,0)
    font_path = os.path.join(abs_path_of_directory, 'certificate_pdf', 'data', 'fonts', 'lucida calligraphy italic.ttf')
    font = ImageFont.truetype(font_path, size)
    subject_font = ImageFont.truetype(font_path, 20)

    name_position = (400, 625) # 1150
    subject_position = (350,770)
    date_position = (330,900)
    duration_position = (830, 700)

    session = create_session()

    for i in range(1276,2000):
        name = sheet.cell(row=i, column=1).value
        subject = sheet.cell(row=i, column=2).value
        date = sheet.cell(row=i, column=3).value
        duration = "15 Days Internship" # sheet.cell(row=i, column=4).value
        email = sheet.cell(row=i, column=4).value

        if date:
            if type(date).__name__ == 'datetime':
                date = date.strftime('%d/%m/%Y')
            
            if name and subject and duration and date:
                name = name.title()
                print(f"Generating Certificate for {name} {subject} {date}")
                certificate_template_path = os.path.join(abs_path_of_directory, "certificate_pdf", "data", "certificate.jpg")
                img = Image.open(certificate_template_path)
                image_ed = ImageDraw.Draw(img)
                # image_ed.text((0,0), 'vis', fill=(255,0,0))
                image_ed.text(name_position, name, fill=color, font=font)
                image_ed.text(subject_position, subject, color, font=subject_font)
                image_ed.text(date_position, date, color, font=font)
                image_ed.text(duration_position, duration, color, font=font)
                certificate_path = os.path.join(abs_path_of_directory, 'certificate_pdf', 'certificates', f'{name}_certificate.jpeg')
                img.save(certificate_path)

                if email and check_email(email):
                    # sending certificate with email
                    send_mail(email, session, certificate_path, task='send_completion_letter')
            else:
                print(f"Certificate generation failed for: {name} {subject} {date} {duration}")

    destroy_session(session)

def send_completion_letter(filepath):
    workbook = load_workbook(filepath) # 'data/student_details.xlsx'
    sheet = workbook.active

    size = 35 
    color = (0,0,0)
    font = ImageFont.truetype('certificate_pdf/data/fonts/lucida calligraphy italic.ttf',size)
    subject_font = ImageFont.truetype('certificate_pdf/data/fonts/lucida calligraphy italic.ttf',20)

    date_position = (110,470)
    name_position = (440, 740)
    subject_position = (500,800)

    session = create_session()

    for i in range(2,2000):
        name = sheet.cell(row=i, column=1).value
        subject = sheet.cell(row=i, column=2).value
        date = sheet.cell(row=i, column=3).value
        email = sheet.cell(row=i, column=4).value

        if date:
            if type(date).__name__ == 'datetime':
                date = date.strftime('%d/%m/%Y')
            
            if name and subject and date and email:
                name = name.title()
                print(f"Generating Completion Letter for {name} {subject} {date}")
                img = Image.open("certificate_pdf/data/completion_letter_template.jpg")
                image_ed = ImageDraw.Draw(img)

                image_ed.text(date_position, date, color, font=font)
                image_ed.text(name_position, name, fill=color, font=font)
                image_ed.text(subject_position, subject, color, font=subject_font)
                
                # image_ed.text(duration_position, duration, color, font=font)
                certificate_path = f'certificate_pdf/completion_letter/{name}_completion_letter.jpeg'
                print(f"certificate_path: {certificate_path}")
                img.save(certificate_path)

                # sending certificate with email
                if email and check_email(email):
                    send_mail(email, session, certificate_path, task='send_completion_letter')

    destroy_session(session)
