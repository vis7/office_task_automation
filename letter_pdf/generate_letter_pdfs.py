"""
This file take excel containing list of students and generate pdf letter for them 
from given template.
"""
###############################################################
# Adding absolute path to the root directory for using absolute path
###############################################################
import os
import sys

abs_path_of_directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.sys.path.append(abs_path_of_directory)
################################################################
import platform
from docx import Document
from openpyxl import load_workbook
from letter_pdf.utils import send_mail, create_session, destroy_session, convert_to_pdf, check_email

os_type = platform.system()
sample_doc_dir = os.path.join('letter_pdf', 'docs')

session = create_session()

# create folders if not exist
def send_confirmation_letter(excel_file_path):
    # getting name from excel sheet
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # generate letter in doc format
    for i in range(705, 1000):
        name = sheet.cell(row=i, column=1).value
        subject = sheet.cell(row=i, column=2).value
        date = sheet.cell(row=i, column=3).value
        email = sheet.cell(row=i, column=4).value

        # refine email
        if email:
            email = email.strip()

        # reformate date to proper format
        if date: # if date or any field is missing then skip the row
            if type(date).__name__ == 'datetime':
                date = date.strftime('%d/%m/%Y')

            print(f"Generating certificate for: {name} {subject} {date}")

            #######################################
            # code for replacing name in doc
            #######################################
            #open the document
            letter_template_path = os.path.join('letter_pdf', 'data', 'joining_letter_template.docx')
            doc=Document(letter_template_path)

            for p in doc.paragraphs:
                inline = p.runs
                for i in range(len(inline)):
                    text = inline[i].text
                    text=text.replace('Name',name)
                    text=text.replace('Subject', subject)
                    text=text.replace('Date', date)
                    inline[i].text = text
            
            # save changed document
            letter_doc_path = os.path.join('letter_pdf', 'docs', f'Internship Confirmation {name}.docx')
            doc.save(letter_doc_path)

            letter_pdf_path = os.path.join('letter_pdf', 'docs', f'Internship Confirmation {name}.pdf')
            convert_to_pdf(letter_doc_path, letter_pdf_path)

            if email and check_email(email):
                # sending mail
                send_mail(email, session, letter_pdf_path)

    destroy_session(session)

    # remove docs
    for file_name in os.listdir(sample_doc_dir):
        os.remove(os.path.join(sample_doc_dir, file_name))
